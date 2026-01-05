import csv, io, re
from django.contrib import messages
from django.contrib.auth import get_user_model
from user.decorators import requiere_rol
from user.models import Rol
from turnos.forms import CargaCSVForm
from turnos.models import Estudiante, RelacionRepresentacion, FuenteRelacion
from django.http import HttpResponse

from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from .models import PerfilDocente, DisponibilidadSemanal, ExcepcionDisponibilidad, TipoExcepcion, FeriadoInstitucional, Cita, ComentarioCita
from .forms import PerfilDocenteForm, DisponibilidadSemanalForm, ExcepcionDisponibilidadForm, BloqueoMasivoForm

from turnos.forms import CargaCSVDocentesForm
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from django.utils.timezone import localtime

from datetime import timedelta, datetime
from django.db import transaction
from django.utils import timezone
from user.decorators import requiere_roles
from user.utils import *



User = get_user_model()

@requiere_roles("Administrador", "DocenteAdministrador")
def cargar_estudiantes(request):
    inst = request.user.institucion
    if request.method == "POST":
        form = CargaCSVForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            try:
                data = archivo.read().decode("utf-8")
            except UnicodeDecodeError:
                messages.error(request, "El archivo debe estar en UTF-8.")
                return redirect("cargar_estudiantes")

            reader = csv.DictReader(io.StringIO(data))
            total_est = total_rel = 0
            rol_rep, _ = Rol.objects.get_or_create(nombre="Representante", defaults={"descripcion":"Padre/Madre/Apoderado"})

            for row in reader:
                cedula = (row.get("cedula") or "").strip()
                nombre = (row.get("nombre") or "").strip()
                curso = (row.get("curso") or "").strip()

                if not cedula or not nombre:
                    continue

                est, created = Estudiante.objects.get_or_create(
                    cedula=cedula,
                    institucion=inst,
                    defaults={"nombre": nombre, "curso": curso}
                )

                if not created:
                    cambios = []
                    if est.nombre != nombre:
                        est.nombre = nombre; cambios.append("nombre")
                    if curso and est.curso != curso:
                        est.curso = curso; cambios.append("curso")
                    if cambios:
                        est.save(update_fields=cambios)
                total_est += 1

                rep_cedula = (row.get("representante_cedula") or "").strip()
                rep_email = (row.get("representante_email") or "").strip().lower()

                if not rep_cedula and not rep_email:
                    # Solo carga estudiante
                    continue

                rep = None
                if rep_cedula:
                    rep = User.objects.filter(cedula=rep_cedula, institucion=inst).first()
                if not rep and rep_email:
                    rep = User.objects.filter(email=rep_email, Institucion=inst).first()

                if not rep:
                    if rep_cedula:
                        rep = User.objects.create_user(
                            username=rep_cedula,
                            cedula=rep_cedula,
                            institucion=inst,
                            email=rep_email or None,
                            password="12345678"
                        )

                    else:
                        rep = User.objects.create_user(
                            username=rep_email, institucion=inst, email=rep_email, password="12345678"
                        )

                actualizar = []
                if rep_cedula and not getattr(rep, "cedula", None):
                    rep.cedula = rep_cedula; actualizar.append("cedula")
                if rep_email and not rep.email:
                    rep.email = rep_email; actualizar.append("email")
                if not rep.rol:
                    rep.rol = rol_rep; actualizar.append("rol")
                if actualizar:
                    rep.save(update_fields=actualizar)

                parentesco = (row.get("parentesco") or "").strip()
                verificado = (row.get("verificado") or "0").strip() in ["1", "true", "True"]
                RelacionRepresentacion.objects.update_or_create(
                    estudiante=est,
                    representante=rep,
                    defaults={
                        "parentesco": parentesco,
                        "verificado": verificado,
                        "fuente": FuenteRelacion.IMPORT,
                        "activo": True,
                    },
                )
                total_rel += 1

            messages.success(request, f"Se importaron {total_est} estudiantes y {total_rel} relaciones.")
            return redirect("cargar_estudiantes")
    else:
        form = CargaCSVForm()

    return render(request, "cargar_estudiantes.html", {"form": form})


@requiere_roles("Administrador", "DocenteAdministrador")
def descargar_formato_estudiantes(request):
    contenido = (
        "cedula,nombre,curso,representante_cedula,representante_email,parentesco,verificado\n"
        "0912345678,Ana Pérez,8vo A,1102345678,rep1@demo.com,Madre,1\n"
        "0912345679,Carlos Gómez,8vo A,1109876543,rep2@demo.com,Padre,0\n"
        "0912345680,Sofía Torres,9no B,1103456789,,Tía,0\n"
    )
    response = HttpResponse(contenido, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="formato_estudiantes.csv"'
    return response

@requiere_roles("Administrador", "DocenteAdministrador")
def listar_docentes(request):
    inst = get_institucion_activa(request)
    # Usuarios con rol Docente o candidatos a serlo (filtro rápido por texto)
    q = (request.GET.get("q") or "").strip().lower()
    inst = request.user.institucion
    rol_doc, _ = Rol.objects.get_or_create(nombre="Docente")

    usuarios = User.objects.filter(
        institucion=inst,
    ).filter(
        Q(rol=rol_doc) | Q(perfil_docente__isnull=False)
    ).select_related("perfil_docente","rol")

    if q:
        usuarios = usuarios.filter(Q(email__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))
    usuarios = usuarios.order_by("first_name","last_name","email")
    return render(request, "docentes_listar.html", {"usuarios": usuarios, "q": q})

@requiere_roles("Administrador", "DocenteAdministrador")
def editar_perfil_docente(request, user_id):
    inst = request.user.institucion

    rol_doc, _ = Rol.objects.get_or_create(nombre="Docente")
    # Solo usuarios de la misma institución
    usuario = get_object_or_404(User, pk=user_id, institucion=inst)

    # PerfilDocente ligado a la misma institución
    perfil, _ = PerfilDocente.objects.get_or_create(
        usuario=usuario,
        defaults={"institucion": inst}
    )

    if not usuario.rol:
        usuario.rol = rol_doc
        usuario.save(update_fields=["rol"])

    if request.method == "POST":
        form = PerfilDocenteForm(request.POST, instance=perfil)
        if form.is_valid():
            perfil = form.save(commit=False)
            perfil.institucion = inst  # reforzamos
            perfil.save()
            messages.success(request, "Perfil de docente guardado.")
            return redirect("gestionar_disponibilidad_docente", docente_id=perfil.id)
    else:
        form = PerfilDocenteForm(instance=perfil)

    return render(request, "docente_editar.html", {
        "usuario": usuario,
        "form": form,
        "perfil": perfil,
    })

@requiere_roles("Administrador", "DocenteAdministrador")
def gestionar_disponibilidad_docente(request, docente_id):
    inst = request.user.institucion

    docente = get_object_or_404(PerfilDocente, pk=docente_id, institucion=inst)

    form_disp = DisponibilidadSemanalForm(request.POST or None)
    form_exc = ExcepcionDisponibilidadForm(request.POST or None)

    if request.method == "POST":
        # Alta disponibilidad semanal
        if "guardar_disp" in request.POST and form_disp.is_valid():
            disp = form_disp.save(commit=False)
            disp.docente = docente
            disp.institucion = inst

            # Regla simple anti-solape de disponibilidad (mismo día)
            solapa = DisponibilidadSemanal.objects.filter(
                docente=docente,
                institucion=inst,
                dia_semana=disp.dia_semana,
                hora_inicio__lt=disp.hora_fin,
                hora_fin__gt=disp.hora_inicio
            ).exists()

            if solapa:
                messages.error(request, "La franja se solapa con otra existente.")
            else:
                disp.full_clean()
                disp.save()
                messages.success(request, "Franja semanal agregada.")
            return redirect("gestionar_disponibilidad_docente", docente_id=docente.id)

        # Alta excepción (bloqueo/extra)
        if "guardar_exc" in request.POST and form_exc.is_valid():
            exc = form_exc.save(commit=False)
            exc.docente = docente
            exc.institucion = inst
            exc.full_clean()
            exc.save()
            messages.success(request, "Excepción guardada.")
            return redirect("gestionar_disponibilidad_docente", docente_id=docente.id)

    disponibilidades = docente.disponibilidades.filter(
        institucion=inst
    ).order_by("dia_semana", "hora_inicio")

    excepciones = docente.excepciones.filter(
        institucion=inst
    ).order_by("-fecha", "hora_inicio")[:30]

    return render(request, "docente_disponibilidad.html", {
        "docente": docente,
        "disponibilidades": disponibilidades,
        "excepciones": excepciones,
        "form_disp": form_disp,
        "form_exc": form_exc,
    })


@requiere_roles("Administrador", "DocenteAdministrador")
def eliminar_disponibilidad(request, disp_id):
    inst = request.user.institucion
    disp = get_object_or_404(DisponibilidadSemanal, pk=disp_id, institucion=inst)
    docente_id = disp.docente.id
    disp.delete()
    messages.info(request, "Franja eliminada.")
    return redirect("gestionar_disponibilidad_docente", docente_id=docente_id)


@requiere_roles("Administrador", "DocenteAdministrador")
def eliminar_excepcion(request, exc_id):
    inst = request.user.institucion
    exc = get_object_or_404(ExcepcionDisponibilidad, pk=exc_id, institucion=inst)
    docente_id = exc.docente.id
    exc.delete()
    messages.info(request, "Excepción eliminada.")
    return redirect("gestionar_disponibilidad_docente", docente_id=docente_id)

ABREV_DIA = {"LUN":0,"MAR":1,"MIE":2,"JUE":3,"VIE":4,"SAB":5,"DOM":6}
RANGO_RE = re.compile(r"^\s*(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})\s*$")

def _parsear_disponibilidad(cadena:str):
    if not cadena: 
        return [], []
    items = [p.strip() for p in cadena.split("|") if p.strip()]
    resultado, errores = [], []
    for seg in items:
        try:
            abrev, rango = seg.split(None, 1)
        except ValueError:
            errores.append(f"Formato inválido: '{seg}'"); 
            continue
        abrev = abrev.upper()
        if abrev not in ABREV_DIA:
            errores.append(f"Día inválido: '{abrev}'"); 
            continue
        m = RANGO_RE.match(rango)
        if not m:
            errores.append(f"Rango inválido: '{rango}' (usa HH:MM-HH:MM)")
            continue
        ini, fin = m.group(1), m.group(2)
        if ini >= fin:
            errores.append(f"Inicio >= fin en '{seg}'")
            continue
        resultado.append((ABREV_DIA[abrev], ini, fin))
    return resultado, errores

@requiere_roles("Administrador", "DocenteAdministrador")
def cargar_docentes(request):
    inst = request.user.institucion  # ← INSTITUCIÓN ACTIVA

    if request.method == "POST":
        form = CargaCSVDocentesForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                data = form.cleaned_data["archivo"].read().decode("utf-8")
            except UnicodeDecodeError:
                messages.error(request, "El archivo debe estar en UTF-8.")
                return redirect("cargar_docentes")

            reader = csv.DictReader(io.StringIO(data))
            rol_doc, _ = Rol.objects.get_or_create(nombre="Docente", defaults={"descripcion":"Docente"})
            creados = actualizados = franjas_creadas = 0

            for row in reader:
                cedula = (row.get("cedula") or "").strip()
                if not cedula:
                    continue

                email = (row.get("email") or "").strip().lower() or None
                nombres = (row.get("nombres") or "").strip()
                apellidos = (row.get("apellidos") or "").strip()
                telefono = (row.get("telefono") or "").strip()
                departamento = (row.get("departamento") or "").strip()

                try:
                    minutos_por_bloque = int(row.get("minutos_por_bloque") or 20)
                except ValueError:
                    minutos_por_bloque = 20

                try:
                    mcd = row.get("maximo_citas_diarias")
                    maximo_citas_diarias = int(mcd) if mcd not in (None,""," ") else None
                except ValueError:
                    maximo_citas_diarias = None

                activo = str(row.get("activo") or "1").strip() in ("1","true","True","TRUE")
                disponibilidad_raw = (row.get("disponibilidad") or "").strip()
                reemplazar = str(row.get("reemplazar_disponibilidad") or "0").strip() in ("1","true","True","TRUE")

                # ---------------------------------------
                # Usuario (username = cedula)
                # ---------------------------------------
                user = User.objects.filter(cedula=cedula, institucion=inst).first()
                if not user:
                    user = User.objects.create_user(
                        username=cedula,
                        cedula=cedula,
                        institucion=inst,  # ← INSTITUCIÓN
                        email=email,
                        password="12345678"
                    )
                    cambios = []
                    if nombres: user.first_name = nombres; cambios.append("first_name")
                    if apellidos: user.last_name = apellidos; cambios.append("last_name")
                    if cambios: user.save(update_fields=cambios)
                    creados += 1
                else:
                    cambios = []
                    if email and not user.email:
                        user.email = email; cambios.append("email")
                    if nombres and user.first_name != nombres:
                        user.first_name = nombres; cambios.append("first_name")
                    if apellidos and user.last_name != apellidos:
                        user.last_name = apellidos; cambios.append("last_name")
                    if cambios:
                        user.save(update_fields=cambios)
                    actualizados += 1

                # Rol Docente
                if not user.rol:
                    user.rol = rol_doc
                    user.save(update_fields=["rol"])

                # ---------------------------------------
                # PerfilDocente
                # ---------------------------------------
                perfil, _ = PerfilDocente.objects.get_or_create(
                    usuario=user,
                    defaults={"institucion": inst}   # ← INSTITUCIÓN
                )

                cambios = []
                if perfil.minutos_por_bloque != minutos_por_bloque:
                    perfil.minutos_por_bloque = minutos_por_bloque; cambios.append("minutos_por_bloque")
                if perfil.maximo_citas_diarias != maximo_citas_diarias:
                    perfil.maximo_citas_diarias = maximo_citas_diarias; cambios.append("maximo_citas_diarias")
                if perfil.departamento != departamento:
                    perfil.departamento = departamento; cambios.append("departamento")
                if perfil.telefono != telefono:
                    perfil.telefono = telefono; cambios.append("telefono")
                if perfil.activo != activo:
                    perfil.activo = activo; cambios.append("activo")

                if cambios:
                    perfil.save(update_fields=cambios)

                # ---------------------------------------
                # Disponibilidad inicial
                # ---------------------------------------
                if disponibilidad_raw:
                    franjas, errs = _parsear_disponibilidad(disponibilidad_raw)
                    if errs:
                        messages.warning(
                            request,
                            f"Fila con cédula {cedula}: {', '.join(errs[:3])}"
                            + (" ..." if len(errs) > 3 else "")
                        )
                    else:
                        if reemplazar:
                            DisponibilidadSemanal.objects.filter(docente=perfil, institucion=inst).delete()

                        for dia, h_ini, h_fin in franjas:
                            existe = DisponibilidadSemanal.objects.filter(
                                docente=perfil,
                                institucion=inst,
                                dia_semana=dia,
                                hora_inicio=h_ini,
                                hora_fin=h_fin
                            ).exists()

                            if not existe:
                                DisponibilidadSemanal.objects.create(
                                    docente=perfil,
                                    institucion=inst,  # ← INSTITUCIÓN
                                    dia_semana=dia,
                                    hora_inicio=h_ini,
                                    hora_fin=h_fin,
                                )
                                franjas_creadas += 1

            messages.success(
                request,
                f"Docentes cargados. Nuevos: {creados}, Actualizados: {actualizados}, Franjas creadas: {franjas_creadas}."
            )
            return redirect("cargar_docentes")
    else:
        form = CargaCSVDocentesForm()

    return render(request, "cargar_docentes.html", {"form": form})

@requiere_roles("Administrador", "DocenteAdministrador")
def formato_docentes(request):
    contenido = (
        "cedula,email,nombres,apellidos,telefono,departamento,minutos_por_bloque,maximo_citas_diarias,activo,disponibilidad,reemplazar_disponibilidad\n"
        "1102345678,doc1@demo.com,María,Lopez,0999999999,Matemática,20,8,1,LUN 08:00-10:00|MAR 09:00-11:00,1\n"
        "1103456789,doc2@demo.com,Juan,Perez,0988888888,Inglés,20,6,1,,0\n"
    )
    resp = HttpResponse(contenido, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="formato_docentes.csv"'
    return resp


def _daterange(d1, d2):
    cur = d1
    while cur <= d2:
        yield cur
        cur += timedelta(days=1)

@requiere_roles("Administrador", "DocenteAdministrador")
@transaction.atomic
def bloqueo_masivo(request):
    inst = request.user.institucion  # ← INSTITUCIÓN

    if request.method == "POST":
        form = BloqueoMasivoForm(request.POST)
        if form.is_valid():
            nombre = form.cleaned_data["nombre"]
            fi = form.cleaned_data["fecha_inicio"]
            ff = form.cleaned_data["fecha_fin"]
            hi = form.cleaned_data.get("hora_inicio")
            hf = form.cleaned_data.get("hora_fin")
            aplicar_a = form.cleaned_data["aplicar_a"]
            depto = (form.cleaned_data.get("departamento") or "").strip()
            reemplazar = form.cleaned_data["reemplazar"]

            # → SOLO docentes de la institución
            docentes = PerfilDocente.objects.filter(activo=True, institucion=inst)

            if aplicar_a == "departamento":
                docentes = docentes.filter(departamento__iexact=depto)

            total_doc = docentes.count()
            if total_doc == 0:
                messages.warning(request, "No hay docentes que coincidan con el filtro.")
                return redirect("bloqueo_masivo")

            # Registrar feriado institucional
            feriado = FeriadoInstitucional.objects.create(
                institucion=inst,  # ← INSTITUCIÓN
                nombre=nombre,
                fecha_inicio=fi,
                fecha_fin=ff,
                hora_inicio=hi,
                hora_fin=hf
            )

            creadas = 0
            for fecha in _daterange(fi, ff):
                for d in docentes:
                    qs = ExcepcionDisponibilidad.objects.filter(
                        docente=d,
                        institucion=inst,  # ← INSTITUCIÓN
                        fecha=fecha,
                        tipo=TipoExcepcion.BLOQUEO
                    )
                    if reemplazar:
                        qs.delete()

                    h_ini = hi or datetime.strptime("00:00", "%H:%M").time()
                    h_fin = hf or datetime.strptime("23:59", "%H:%M").time()

                    if not qs.filter(hora_inicio=h_ini, hora_fin=h_fin).exists():
                        ExcepcionDisponibilidad.objects.create(
                            docente=d,
                            institucion=inst,   # ← INSTITUCIÓN
                            fecha=fecha,
                            hora_inicio=h_ini,
                            hora_fin=h_fin,
                            tipo=TipoExcepcion.BLOQUEO,
                            motivo=nombre
                        )
                        creadas += 1

            messages.success(
                request,
                f"Bloqueo aplicado: {feriado.nombre}. Docentes: {total_doc}. Excepciones creadas: {creadas}."
            )
            return redirect("bloqueo_masivo")
    else:
        form = BloqueoMasivoForm()

    return render(request, "bloqueo_masivo.html", {"form": form})

@requiere_roles("Administrador", "DocenteAdministrador", "Docente")
@require_POST
def cita_aprobar(request, pk):
    inst = request.user.institucion
    cita = get_object_or_404(Cita, pk=pk, institucion=inst)

    if cita.estado != "PROPUESTA":
        messages.error(request, "Solo se pueden aprobar citas en estado PROPUESTA.")
        return redirect("admin_buscar_citas")

    cita.estado = "CONFIRMADA"
    cita.save()

    messages.success(request, "Cita aprobada correctamente.")
    return redirect("admin_buscar_citas")


@requiere_roles("Administrador", "DocenteAdministrador", "Docente")
@require_POST
def cita_rechazar(request, pk):
    inst = request.user.institucion
    cita = get_object_or_404(Cita, pk=pk, institucion=inst)

    if cita.estado != "PROPUESTA":
        messages.error(request, "Solo se pueden rechazar citas en estado PROPUESTA.")
        return redirect("admin_buscar_citas")

    cita.estado = "RECHAZADA"
    cita.save()

    messages.success(request, "Cita rechazada.")
    return redirect("admin_buscar_citas")


import os
from openpyxl.drawing.image import Image as XLImage
from django.conf import settings
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from turnos.forms import FiltroCitasForm

@requiere_roles("Administrador", "DocenteAdministrador")
def exportar_citas_excel(request):
    inst = request.user.institucion
    hoy = timezone.localdate()
    form = FiltroCitasForm(request.GET or None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Citas y Comentarios"
    fila_encabezados = 10

    # =========================
    # LOGOS
    # =========================

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD",  # azul sobrio (puedes cambiar)
        end_color="4F81BD",
        fill_type="solid"
    )
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )


    # Logo institución (si existe)
    ws.merge_cells("A1:B6")
    if inst.logo and os.path.exists(inst.logo.path):
        logo_inst = XLImage(inst.logo.path)
        logo_inst.width = 110
        logo_inst.height = 110
        ws.add_image(logo_inst, "A1")  # anclado a la esquina superior izquierda    # Logo EasyDate (static) En producción
    #logo_easydate_path = os.path.join(
     #   settings.BASE_DIR,"static", "images","img","LogoEasyDateTr.png")
    
    #LOGO PARA PROBAR EN LOCAL
    logo_easydate_path = os.path.join(
        settings.BASE_DIR,"myApp","static", "assets","images","img","LogoEasyDateTr.png")
    
    print(logo_easydate_path)
    # Logo EasyDate
    ws.merge_cells("E1:F6")
    if os.path.exists(logo_easydate_path):
        logo_easy = XLImage(logo_easydate_path)
        logo_easy.width = 110
        logo_easy.height = 110
        ws.add_image(logo_easy, "E1")

    # =========================
    # TÍTULO
    # =========================
    ws.merge_cells("B8:H8")
    ws["B8"] = f"Reporte de Citas - {inst.nombre}"
    ws["B8"].style = "Title"


    # =========================
    # ENCABEZADOS
    # =========================
    headers = [
        "ID Cita",
        "Fecha",
        "Hora inicio",
        "Hora Fin",
        "Docente",
        "Representante",
        "Email del representante",
        "Estudiante",
        "Curso",
        "Estado",
        "Motivo",
        "Autor comentario",
        "Fecha comentario",
        "Comentario",
    ]
    ws.append([])
    ws.append(headers)
    fila_encabezados = ws.max_row

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=fila_encabezados, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = f"A{fila_encabezados + 1}"

    # =========================
    # DATOS
    # =========================
    citas = (
        Cita.objects
        .filter(institucion=inst)
        .select_related("docente__usuario", "representante")
        .prefetch_related("comentarios__autor")
        .order_by("inicio")
    )

    if form.is_valid():
        fecha = form.cleaned_data.get("fecha")
        docente = form.cleaned_data.get("docente")
        estado = form.cleaned_data.get("estado")
        if fecha:
            citas = citas.filter(inicio__date=fecha)
        else:
            desde, hasta = hoy, hoy + timezone.timedelta(days=7)
            citas = citas.filter(inicio__date__range=(desde, hasta))
        if docente:
            citas = citas.filter(docente=docente)
        if estado:
            citas = citas.filter(estado=estado)    

    for cita in citas:
        comentarios = cita.comentarios.all()
        ini = localtime(cita.inicio).replace(tzinfo=None)
        fin = localtime(cita.fin).replace(tzinfo=None)

        if not comentarios:
            ws.append([
                cita.id,
                ini.date(),
                ini.time(),
                fin.time(),
                cita.docente.usuario.get_full_name(),
                cita.representante.get_full_name(),
                cita.representante.email,
                cita.nombre_estudiante,
                cita.curso_estudiante,
                cita.get_estado_display(),
                (cita.motivo or "")[:150],
                "",
                "",
                "",
            ])
            continue

        for com in comentarios:
            ws.append([
                cita.id,
                ini.date(),
                ini.time(),
                fin.time(),
                cita.docente.usuario.get_full_name(),
                cita.representante.get_full_name(),
                cita.representante.email,
                cita.nombre_estudiante,
                cita.curso_estudiante,
                cita.get_estado_display(),                
                (cita.motivo or "")[:150],
                com.autor.get_full_name() or com.autor.username,
                localtime(com.creado_en).replace(tzinfo=None),
                com.texto,
            ])

    # =========================
    # RESPUESTA
    # =========================

    ultima_fila = ws.max_row
    ultima_col = ws.max_column
    
    col_final = get_column_letter(ultima_col)

    ws.auto_filter.ref = f"A{fila_encabezados}:{col_final}{ultima_fila}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_citas_easydate.xlsx"'
    wb.save(response)
    return response
